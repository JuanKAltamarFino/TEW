import pandas as pd
from pathlib import Path
import json
import numpy as np
def createDF(elements,lst_columns):
	return pd.DataFrame(elements,columns=lst_columns)
def writeExcel(xlsx_file,dict_data):
	from openpyxl.utils import get_column_letter
	v_mod_with=3*1.094
	with pd.ExcelWriter(xlsx_file, engine="openpyxl", mode='w') as writer:
		for key in dict_data.keys():
			dict_data[key].to_excel(writer,index=False,sheet_name=key)
			for column in dict_data[key]:
				column_width = np.nanmax([dict_data[key][column].astype(str).map(len).max()+v_mod_with, len(str(column))+v_mod_with])
				col_idx = dict_data[key].columns.get_loc(column)
				#debugging(writer.sheets[key].column_dimensions[get_column_letter(col_idx+1)])
				writer.sheets[key].column_dimensions[get_column_letter(col_idx+1)].width = column_width
def debugging(data):
    printTypeAndContent(data)
    raise Exception("JuanK is debugging")
def printTypeAndContent(data):
    print(f'printTypeAndContent\n{type(data)}\n{data}')
def readJsonFile(v_folder,v_file_name):
	dict_=dict()
	json_file = Path(v_folder, v_file_name)
	with open(json_file) as json_file:
		dict_ = json.load(json_file)
	return dict_
def writeJson(dict_,folder,file_name):
	v_path=Path(folder, file_name)
	with open(v_path, 'w') as f:
		json.dump(dict_, f, indent=2)
def dictToDictOfDataFrame(dict_):
	dict_df={}
	for key in dict_.keys():
		df = dictToDataFrameByKey(dict_,key)
		dict_df.update({key:df})
	return dict_df
def dictToDataFrameByKey(dict_,key):
	df = pd.DataFrame.from_dict(dict_[key])
	return df
def getLen(obj):
	val = obj
	if val is None:
		return 0
	else:
		if isinstance(obj, pd.DataFrame):
			return len(val.values)
		else:
			return len(val)
def saveObjectForTesting(file_name,object):
	import pickle
	# Step 2
	fullRoute=generateFullRouteTestFile(file_name)
	with open(fullRoute, 'wb') as file:
	  # Step 3
	  pickle.dump(object, file)
def generateFullRouteTestFile(file_name):
	createTestFolderIfNotExist()
	folder=getTestObjectsFolder()
	return folder+"/"+file_name
def saveDFObjectForTesting(file_name,object):
	if object is None:
		return
	fullRoute=generateFullRouteTestFile(file_name)
	object.to_pickle(fullRoute)
def readObjectForTesting(file_name):
	# Step 1
	import pickle
	# Step 2
	file_name=generateFullRouteTestFile(file_name)
	with open(file_name, 'rb') as file:
		# Step 3
		object = pickle.load(file)
		# After config_dictionary is read from file
		return object
def readDFObjectForTesting(file_name):
	file_name=generateFullRouteTestFile(file_name)
	unpickled_df = pd.read_pickle(file_name) 
	return unpickled_df
def readBasicConfigValuesJson():
	v_folder="./"
	v_file_name="BasicConfigValues.json"
	dict_basic_values=readJsonFile(v_folder,v_file_name)
	return dict_basic_values
def isSublistInList(sublist,test_list):
	c=0
	res=False
	for i in sublist:
		if i in test_list:
			c+=1
	if(c==len(sublist)):
		res=True
	return res
def filterListElementsByRegex(patterns,list_):
	import re
	result_list = []
	for pattern in patterns:
		filter_list = list(filter(lambda x: re.search(pattern, x), list_))
		if len(filter_list) > 0:
			result_list.extend(filter_list)
	return result_list
def findingDuplicateItemsInList(list_):
	duplicates = [item for item in list_ if list_.count(item) > 1]
	unique_duplicates = list(set(duplicates))
	return unique_duplicates
def createTestFolderIfNotExist():
	import os
	test_folder=getTestObjectsFolder()
	if not os.path.exists(test_folder):
		os.mkdir(test_folder)
def getTestObjectsFolder():
	return './TestObjects'